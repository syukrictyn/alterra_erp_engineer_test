from odoo import models, fields, api
import logging
import base64
from io import BytesIO

_logger = logging.getLogger(__name__)

try:
    # optional: if queue_job installed, we'll use decorator
    from odoo.addons.queue_job.job import job
    QUEUE_JOB_AVAILABLE = True
except Exception:
    QUEUE_JOB_AVAILABLE = False

class EmployeeImportJob(models.Model):
    _name = "employee.import.job"
    _description = "Employee Import Job"

    name = fields.Char(default="Employee Import")
    user_id = fields.Many2one('res.users', default=lambda self: self.env.uid)
    file = fields.Binary("File (xlsx)", required=True)
    file_name = fields.Char("File Name")
    state = fields.Selection([
        ('draft', 'Draft'),
        ('pending', 'Pending'),
        ('running', 'Running'),
        ('done', 'Done'),
        ('failed', 'Failed'),
    ], default='draft')
    total = fields.Integer()
    processed = fields.Integer()
    errors = fields.Text()

    def action_start(self):
        """Mark pending and enqueue processing."""
        for rec in self:
            rec.state = 'pending'
            if QUEUE_JOB_AVAILABLE:
                rec._enqueue_process()
        return True

    if QUEUE_JOB_AVAILABLE:
        @job(default_channel='root.job')
        def _enqueue_process(self):
            return self.with_env(self.env).process_file()

    def process_file(self):
        """Main worker: read xlsx (openpyxl in read_only) and create hr.employee records."""
        try:
            from openpyxl import load_workbook
        except Exception:
            _logger.exception("openpyxl not installed")
            for r in self:
                r.state = 'failed'
                r.errors = 'openpyxl not installed'
            return

        for job in self:
            job.state = 'running'
            job.processed = 0
            job.total = 0
            job.errors = False

            try:
                data = base64.b64decode(job.file)
                stream = BytesIO(data)
                wb = load_workbook(filename=stream, read_only=True)
                ws = wb.active
            except Exception as e:
                job.state = 'failed'
                job.errors = "Failed to read file: %s" % (e,)
                _logger.exception("Failed to open xlsx")
                job._notify_user()
                continue

            rows = ws.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                job.state = 'failed'
                job.errors = "Empty file"
                job._notify_user()
                continue

            header_map = { (h or '').strip().lower(): i for i,h in enumerate(headers) }
            created = 0
            errors = []

            for row_idx, row in enumerate(rows, start=2):
                try:
                    name = row[ header_map.get('name') ] if header_map.get('name') is not None else None
                    work_email = row[ header_map.get('work_email') ] if header_map.get('work_email') is not None else None
                    ident = row[ header_map.get('identification_id') ] if header_map.get('identification_id') is not None else None
                    work_phone = row[ header_map.get('work_phone') ] if header_map.get('work_phone') is not None else None
                except Exception as e:
                    errors.append(f"Row {row_idx}: mapping error - {e}")
                    _logger.exception("Row mapping error")
                    continue

                if not name:
                    errors.append(f"Row {row_idx}: missing name")
                    continue

                # check duplicates by identification_id or email
                domain = []
                if ident:
                    domain.append(('identification_id', '=', str(ident)))
                if work_email:
                    domain.append(('work_email', '=', work_email))

                exists = None
                if domain:
                    exists = self.env['hr.employee'].sudo().search(domain, limit=1)

                if exists:
                    errors.append(f"Row {row_idx}: duplicate (email or id)")
                    continue

                try:
                    vals = {
                        'name': name,
                        'work_email': work_email or False,
                        'work_phone': work_phone or False,
                        'identification_id': str(ident) if ident else False,
                    }
                    self.env['hr.employee'].sudo().create(vals)
                    created += 1
                except Exception as e:
                    _logger.exception("Error creating employee row %s", row_idx)
                    errors.append(f"Row {row_idx}: create error - {e}")

            job.total = created + len(errors)
            job.processed = created
            job.errors = "\n".join(errors) if errors else False
            job.state = 'done' if not errors else 'failed'
            job._notify_user()

    def _notify_user(self):
        """Send a brief email to job.user_id (if email configured).
           We create mail.mail directly to avoid xml model reference problems.
        """
        for job in self:
            try:
                email_to = job.user_id.email or False
                subject = f"Employee Import: {job.name} - {job.state}"
                body = f"""
                <p>Hi {job.user_id.name},</p>
                <p>Import job <strong>{job.name}</strong> selesai dengan status: <strong>{job.state}</strong>.</p>
                <p>Processed: {job.processed} / {job.total}</p>
                <p>Errors (if any):</p>
                <pre>{job.errors or 'No errors'}</pre>
                """
                if not email_to:
                    _logger.info("No email for user %s, skipping notification", job.user_id.id)
                    return
                mail_vals = {
                    'subject': subject,
                    'body_html': body,
                    'email_to': email_to,
                }
                self.env['mail.mail'].sudo().create(mail_vals).send()
            except Exception:
                _logger.exception("Failed to send import result email for job %s", job.id)
